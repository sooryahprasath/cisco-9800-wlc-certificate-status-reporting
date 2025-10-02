#!/usr/bin/env python3

"""
Cisco 9800 Series WLC Certificate Status Reporter
A secure, multi-threaded Python application for collecting and reporting 
certificate status information from Cisco 9800 Series Wireless LAN Controllers.
"""

import json
import os
import sys
import time
import threading
import getpass
import hashlib
import secrets
import base64
import ctypes
import gc
import re
from datetime import datetime, timedelta
from dateutil import parser as date_parser
import pytz
from netmiko import ConnectHandler
import concurrent.futures
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from dotenv import load_dotenv
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from ping3 import ping

# Load environment variables
load_dotenv()

# Configuration from environment
WLC_DATA_FILE = os.getenv('WLC_DATA_FILE', "wlc.json")
MAX_THREADS = int(os.getenv('MAX_THREADS', '10'))
COMMAND_TIMEOUT = int(os.getenv('COMMAND_TIMEOUT', '180'))
CONNECTION_TIMEOUT = int(os.getenv('CONNECTION_TIMEOUT', '180'))
PING_TIMEOUT = int(os.getenv('PING_TIMEOUT', '5'))
MAX_RETRIES = int(os.getenv('MAX_RETRIES', '3'))

# Directories
LOG_DIR = os.getenv('LOG_DIR', "logs")
SYSTEM_LOGS_DIR = os.getenv('SYSTEM_LOGS_DIR', "logs")
os.makedirs(LOG_DIR, exist_ok=True)
os.makedirs(SYSTEM_LOGS_DIR, exist_ok=True)

# Email configuration
SMTP_SERVER = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
SMTP_PORT = int(os.getenv('SMTP_PORT', '587'))
SENDER_EMAIL = os.getenv('SENDER_EMAIL')
SENDER_PASSWORD = os.getenv('SENDER_PASSWORD')
DEFAULT_RECIPIENTS = os.getenv('DEFAULT_RECIPIENTS', '').split(',') if os.getenv('DEFAULT_RECIPIENTS') else []

# Timezone configuration
TIMEZONE = os.getenv('REPORT_TIMEZONE', 'UTC')
TZ = pytz.timezone(TIMEZONE)

# Log files
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
LOG_FILE = os.path.join(SYSTEM_LOGS_DIR, f"wlc_cert_check_{timestamp}.log")
ERROR_LOG_FILE = os.path.join(SYSTEM_LOGS_DIR, "wlc_cert_errors.log")

# Thread-safe results storage
results = []
results_lock = threading.Lock()

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(SYSTEM_LOGS_DIR, 'wlc_cert_checker.log')),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Separate error logger
error_logger = logging.getLogger('errors')
error_handler = logging.FileHandler(ERROR_LOG_FILE)
error_handler.setLevel(logging.ERROR)
error_formatter = logging.Formatter('%(asctime)s - ERROR - %(message)s')
error_handler.setFormatter(error_formatter)
error_logger.addHandler(error_handler)
error_logger.setLevel(logging.ERROR)


class SecurePassword:
    """Secure password container with memory protection"""
    
    def __init__(self, password: str):
        self.salt = secrets.token_bytes(32)
        self.key = self._derive_key(password, self.salt)
        self.fernet = Fernet(self.key)
        self.encrypted_password = self.fernet.encrypt(password.encode())
        password = "0" * len(password)
        self._memory_addresses = set()
        
    def _derive_key(self, password: str, salt: bytes) -> bytes:
        """Derive encryption key using PBKDF2"""
        kdf = PBKDF2HMAC(
            algorithm=hashes.SHA256(),
            length=32,
            salt=salt,
            iterations=100000,
        )
        key = base64.urlsafe_b64encode(kdf.derive(password.encode()))
        return key
    
    def get_password(self) -> str:
        """Decrypt and return password"""
        try:
            decrypted = self.fernet.decrypt(self.encrypted_password)
            password = decrypted.decode()
            self._memory_addresses.add(id(password))
            return password
        except Exception:
            raise ValueError("Failed to decrypt password")
    
    def clear_memory(self):
        """Clear password from memory"""
        try:
            if hasattr(self, 'encrypted_password'):
                self.encrypted_password = b"0" * len(self.encrypted_password)
            if hasattr(self, 'key'):
                self.key = b"0" * len(self.key)
            for addr in self._memory_addresses:
                try:
                    ctypes.memset(addr, 0, 64)
                except:
                    pass
        except Exception:
            pass
    
    def __del__(self):
        self.clear_memory()


class SecureWLCSession:
    """Secure session management for WLC operations"""
    
    def __init__(self, session_timeout=1800):
        self.session_timeout = session_timeout
        self.username = None
        self.password = None
        self.last_access = None
        self.session_lock = threading.Lock()
        self.cleanup_timer = None
        
    def authenticate(self) -> bool:
        """Authenticate user and establish secure session"""
        with self.session_lock:
            try:
                print("\n" + "="*60)
                print("WLC Authentication Required")
                print("="*60)
                print("Session will timeout after 30 minutes of inactivity\n")
                
                self.username = input("Enter WLC username: ").strip()
                if not self.username:
                    raise ValueError("Username cannot be empty")
                
                raw_password = getpass.getpass("Enter WLC password: ")
                if not raw_password:
                    raise ValueError("Password cannot be empty")
                
                self.password = SecurePassword(raw_password)
                raw_password = "0" * len(raw_password)
                del raw_password
                
                self.last_access = time.time()
                
                temp_password = self.password.get_password()
                password_hash = hashlib.sha256(temp_password.encode()).hexdigest()[:8]
                temp_password = "0" * len(temp_password)
                del temp_password
                
                print(f"\nAuthentication successful (Session ID: {password_hash})")
                print("Credentials encrypted in memory with AES-128\n")
                
                self._start_cleanup_timer()
                return True
                
            except KeyboardInterrupt:
                print("\n\nAuthentication cancelled by user")
                return False
            except Exception as e:
                print(f"\nAuthentication failed: {e}")
                return False
    
    def get_credentials(self) -> tuple[str, str]:
        """Get credentials if session is valid"""
        with self.session_lock:
            current_time = time.time()
            
            if self.last_access and (current_time - self.last_access > self.session_timeout):
                print("\nSession expired - please authenticate again")
                self._clear_session()
                raise ValueError("Session expired")
            
            if not self.password or not self.username:
                raise ValueError("Not authenticated")
            
            self.last_access = current_time
            password = self.password.get_password()
            return self.username, password
    
    def create_wlc_connection(self, hostname: str, ip: str, thread_id: int) -> dict:
        """Create WLC connection configuration"""
        username, password = self.get_credentials()
        
        connection_config = {
            'device_type': 'cisco_wlc',
            'host': ip,
            'username': username,
            'password': password,
            'timeout': CONNECTION_TIMEOUT,
            'global_delay_factor': 3,
            'session_log': os.path.join(SYSTEM_LOGS_DIR, f'wlc_cert_session_{hostname}_{thread_id}.log'),
            'keepalive': 30,
            'conn_timeout': 30,
            'auth_timeout': 30,
            'banner_timeout': 30,
        }
        
        return connection_config
    
    def _start_cleanup_timer(self):
        """Start automatic session cleanup timer"""
        def cleanup_worker():
            while self.password:
                time.sleep(300)
                
                with self.session_lock:
                    if self.last_access:
                        inactive_time = time.time() - self.last_access
                        
                        if inactive_time > self.session_timeout:
                            print("\nSession expired due to inactivity")
                            self._clear_session()
                            break
        
        if self.cleanup_timer:
            self.cleanup_timer = None
            
        self.cleanup_timer = threading.Thread(target=cleanup_worker, daemon=True)
        self.cleanup_timer.start()
    
    def _clear_session(self):
        """Clear session data securely"""
        try:
            if self.password:
                self.password.clear_memory()
                self.password = None
            
            if self.username:
                self.username = "0" * len(self.username)
                self.username = None
                
            self.last_access = None
        except Exception:
            pass
    
    def is_authenticated(self) -> bool:
        """Check if session is authenticated and valid"""
        with self.session_lock:
            if not self.password or not self.last_access:
                return False
            
            inactive_time = time.time() - self.last_access
            return inactive_time <= self.session_timeout
    
    def __del__(self):
        self._clear_session()


def parse_certificates(output_text):
    """
    Parse certificate information from 'show crypto pki certificates' output
    Returns a list of certificate dictionaries
    """
    certificates = []
    
    # Split by certificate sections
    cert_sections = re.split(r'\n(?=(?:Certificate|CA Certificate|Router Self-Signed Certificate)\s*\n)', output_text)
    
    for section in cert_sections:
        if not section.strip():
            continue
            
        cert = {}
        
        # Determine certificate type
        if section.startswith('CA Certificate'):
            cert['cert_type'] = 'CA Certificate'
        elif section.startswith('Router Self-Signed Certificate'):
            cert['cert_type'] = 'Router Self-Signed Certificate'
        elif section.startswith('Certificate'):
            cert['cert_type'] = 'Certificate'
        else:
            continue
        
        # Extract Status
        status_match = re.search(r'Status:\s*(.+)', section)
        cert['status'] = status_match.group(1).strip() if status_match else 'Unknown'
        
        # Extract Certificate Name (Priority: Name field, then cn=, then Subject)
        name_match = re.search(r'Name:\s*(.+)', section)
        if name_match:
            cert['name'] = name_match.group(1).strip()
        else:
            cn_match = re.search(r'cn=([^,\n]+)', section)
            if cn_match:
                cert['name'] = cn_match.group(1).strip()
            else:
                subject_match = re.search(r'Subject:.*?\n\s+(.+)', section, re.DOTALL)
                if subject_match:
                    first_line = subject_match.group(1).split('\n')[0].strip()
                    cert['name'] = first_line
                else:
                    cert['name'] = 'Unknown'
        
        # Extract Start Date
        start_match = re.search(r'start date:\s*(.+)', section)
        cert['start_date'] = start_match.group(1).strip() if start_match else 'Unknown'
        
        # Extract End Date
        end_match = re.search(r'end\s+date:\s*(.+)', section)
        cert['end_date'] = end_match.group(1).strip() if end_match else 'Unknown'
        
        # Extract Associated Trustpoints
        trustpoint_match = re.search(r'Associated Trustpoints:\s*(.+)', section)
        cert['trustpoints'] = trustpoint_match.group(1).strip() if trustpoint_match else 'None'
        
        # Extract Storage
        storage_match = re.search(r'Storage:\s*(.+)', section)
        if storage_match:
            cert['storage'] = storage_match.group(1).strip()
        else:
            cert['storage'] = cert['name']
        
        certificates.append(cert)
    
    return certificates


def calculate_cert_remarks(end_date_str):
    """
    Calculate remarks based on certificate expiry
    Returns: (remarks_text, expiry_date_obj)
    """
    try:
        end_date = date_parser.parse(end_date_str)
        
        if end_date.tzinfo is None:
            end_date = pytz.UTC.localize(end_date)
        
        now = datetime.now(pytz.UTC)
        time_diff = end_date - now
        
        if time_diff.total_seconds() < 0:
            return "EXPIRED", end_date
        
        days_remaining = time_diff.days
        
        if days_remaining < 365:
            if days_remaining < 30:
                time_str = f"{days_remaining} days remaining"
            else:
                months = days_remaining // 30
                time_str = f"{months} months remaining"
            
            renewal_date = end_date.astimezone(TZ)
            renewal_str = renewal_date.strftime('%d-%m-%Y')
            
            return f"{time_str}, renewal due: {renewal_str}", end_date
        else:
            return "OK", end_date
            
    except Exception as e:
        logger.warning(f"Could not parse date '{end_date_str}': {e}")
        return "UNKNOWN", None


def ping_wlc(ip_address, timeout=PING_TIMEOUT):
    """Ping WLC to check connectivity"""
    try:
        response = ping(ip_address, timeout=timeout)
        return response is not None
    except:
        return False


def connect_to_wlc(hostname, ip_address, secure_session, commands, thread_id=0):
    """Connect to WLC and execute commands"""
    try:
        logger.info(f"[Thread-{thread_id}] Checking connectivity to {hostname} ({ip_address})...")
        if not ping_wlc(ip_address):
            error_msg = f"WLC not reachable (ping failed): {hostname} ({ip_address})"
            logger.error(f"[Thread-{thread_id}] {error_msg}")
            error_logger.error(f"{hostname} - Ping failed to {ip_address}")
            return False, error_msg
        
        wlc_device = secure_session.create_wlc_connection(hostname, ip_address, thread_id)
        
        logger.info(f"[Thread-{thread_id}] Connecting to {hostname} ({ip_address})...")
        net_connect = ConnectHandler(**wlc_device)
        
        wlc_device['password'] = "0" * len(wlc_device['password'])
        del wlc_device
        
        logger.info(f"[Thread-{thread_id}] Connection successful to {hostname}")
        
        # Disable pagination
        try:
            net_connect.send_command("terminal length 0", expect_string=r"#")
        except Exception as e:
            logger.warning(f"[Thread-{thread_id}] Could not disable pagination: {str(e)}")
        
        # Execute commands
        all_output = ""
        
        for command in commands:
            logger.info(f"[Thread-{thread_id}] Running command: '{command}'...")
            
            try:
                output = net_connect.send_command(
                    command,
                    delay_factor=20,
                    max_loops=2000,
                    read_timeout=COMMAND_TIMEOUT
                )
                logger.info(f"[Thread-{thread_id}] Command completed successfully")
                all_output += output
            except Exception as cmd_error:
                logger.warning(f"[Thread-{thread_id}] Command timed out")
                error_logger.error(f"{hostname} - Command timeout: {str(cmd_error)}")
                all_output = f"[Command timed out: {str(cmd_error)}]"
        
        # Disconnect
        try:
            net_connect.disconnect()
            logger.info(f"[Thread-{thread_id}] Connection closed")
        except:
            logger.warning(f"[Thread-{thread_id}] Could not properly close connection")
        
        return True, all_output
        
    except Exception as e:
        error_msg = f"Error connecting to {hostname} ({ip_address}): {str(e)}"
        logger.error(f"[Thread-{thread_id}] {error_msg}")
        error_logger.error(f"{hostname} - Connection error: {str(e)}")
        return False, error_msg


def process_wlc(wlc_data, secure_session, commands, thread_id):
    """Process a single WLC"""
    hostname = wlc_data["HOSTNAME"]
    ip = wlc_data["IP ADDRESS"]
    
    logger.info(f"[Thread-{thread_id}] Processing WLC: {hostname}")
    
    success = False
    output = ""
    
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            success, output = connect_to_wlc(hostname, ip, secure_session, commands, thread_id)
            if success:
                break
            elif "Session expired" in output:
                break
            elif attempt < MAX_RETRIES:
                logger.info(f"[Thread-{thread_id}] Retrying (attempt {attempt + 1}/{MAX_RETRIES})...")
                time.sleep(10)
        except Exception as e:
            output = f"Connection attempt {attempt} failed: {str(e)}"
            logger.error(f"[Thread-{thread_id}] {output}")
            if attempt == MAX_RETRIES:
                break
    
    result = {
        'hostname': hostname,
        'ip': ip,
        'success': success,
        'output': output,
        'thread_id': thread_id,
        'site_id': wlc_data.get('SITE ID', 'Unknown'),
        'region': wlc_data.get('REGION', 'Unknown'),
        'city': wlc_data.get('CITY', 'Unknown'),
        'model': wlc_data.get('MODEL', 'Unknown'),
        'timezone': wlc_data.get('TIMEZONE', 'Unknown')
    }
    
    with results_lock:
        results.append(result)
    
    return result


def create_excel_report(certificate_data, output_filename):
    """Create Excel report with certificate data"""
    logger.info(f"Creating Excel report: {output_filename}")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Certificate Status"
    
    now_tz = datetime.now(TZ)
    report_date = now_tz.strftime('%d-%m-%Y')
    report_time = now_tz.strftime('%H:%M')
    
    headers = [
        'REPORT DATE',
        'REPORT TIME',
        'REGION',
        'SITE ID',
        'WLC HOSTNAME',
        'WLC IP ADDRESS',
        'CERT TYPE',
        'CERT NAME',
        'START DATE',
        'END DATE',
        'ASSOCIATED TRUSTPOINTS',
        'STATUS',
        'STORAGE',
        'REMARKS'
    ]
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    row_num = 2
    for cert_entry in certificate_data:
        ws.cell(row=row_num, column=1, value=report_date)
        ws.cell(row=row_num, column=2, value=f"{report_time} {TIMEZONE}")
        ws.cell(row=row_num, column=3, value=cert_entry['region'])
        ws.cell(row=row_num, column=4, value=cert_entry['site_id'])
        ws.cell(row=row_num, column=5, value=cert_entry['hostname'])
        ws.cell(row=row_num, column=6, value=cert_entry['ip'])
        ws.cell(row=row_num, column=7, value=cert_entry['cert_type'])
        ws.cell(row=row_num, column=8, value=cert_entry['name'])
        ws.cell(row=row_num, column=9, value=cert_entry['start_date'])
        ws.cell(row=row_num, column=10, value=cert_entry['end_date'])
        ws.cell(row=row_num, column=11, value=cert_entry['trustpoints'])
        ws.cell(row=row_num, column=12, value=cert_entry['status'])
        ws.cell(row=row_num, column=13, value=cert_entry['storage'])
        ws.cell(row=row_num, column=14, value=cert_entry['remarks'])
        
        # Apply formatting
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = cell_alignment
            cell.border = thin_border
            
            # Color code remarks
            if cert_entry['remarks'] == 'EXPIRED':
                cell = ws.cell(row=row_num, column=14)
                cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            elif 'remaining' in cert_entry['remarks']:
                cell = ws.cell(row=row_num, column=14)
                cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                cell.font = Font(bold=True)
        
        row_num += 1
    
    # Column widths
    column_widths = {
        'A': 15, 'B': 18, 'C': 12, 'D': 12, 'E': 20, 'F': 18,
        'G': 25, 'H': 35, 'I': 30, 'J': 30, 'K': 35,
        'L': 12, 'M': 40, 'N': 45
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    wb.save(output_filename)
    logger.info(f"Excel report created: {output_filename}")
    
    return output_filename


def send_email(report_file, subject=None, recipients=None, summary_data=None):
    """Send email with report attachment"""
    try:
        send_to = recipients or DEFAULT_RECIPIENTS
        if not send_to:
            logger.warning("No email recipients configured")
            return False
        
        now_tz = datetime.now(TZ)
        subject = subject or f"WLC Certificate Status Report - {now_tz.strftime('%d %b %Y').upper()}"
        
        body = f"""Hello,

Attached is the Cisco WLC Certificate Status Report generated on {now_tz.strftime('%d %B %Y')} at {now_tz.strftime('%H:%M')} {TIMEZONE}.

EXECUTION SUMMARY:
==================
"""
        
        if summary_data:
            body += f"""Total WLCs Processed: {summary_data['total_wlcs']}
Successful Connections: {summary_data['successful']}
Failed Connections: {summary_data['failed']}
Total Certificates Found: {summary_data['total_certs']}
Expired Certificates: {summary_data['expired_certs']}
Certificates Expiring Soon: {summary_data['expiring_soon']}
Execution Time: {summary_data['execution_time']:.2f} seconds ({summary_data['execution_time']/60:.1f} minutes)

"""
            if summary_data.get('failures'):
                body += "FAILURE DETAILS:\n"
                for failure in summary_data['failures']:
                    body += f"- {failure['hostname']} ({failure['ip']}): {failure['reason']}\n"

        body += f"""
This report contains certificate status information from all WLCs in the environment.

For detailed error logs, check the wlc_cert_errors.log file.

Regards,
WLC Certificate Monitoring System
"""

        msg = MIMEMultipart()
        msg['From'] = SENDER_EMAIL
        msg['To'] = ", ".join(send_to)
        msg['Date'] = formatdate(localtime=True)
        msg['Subject'] = subject
        msg.attach(MIMEText(body))
        
        # Add report attachment
        with open(report_file, 'rb') as f:
            attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            attachment.set_payload(f.read())
            encoders.encode_base64(attachment)
            attachment.add_header('Content-Disposition', 
                                 f'attachment; filename="{os.path.basename(report_file)}"')
            msg.attach(attachment)
        
        # Add error log if exists
        if os.path.exists(ERROR_LOG_FILE) and os.path.getsize(ERROR_LOG_FILE) > 0:
            with open(ERROR_LOG_FILE, 'rb') as f:
                error_attachment = MIMEBase('text', 'plain')
                error_attachment.set_payload(f.read())
                encoders.encode_base64(error_attachment)
                error_attachment.add_header('Content-Disposition', 
                                           'attachment; filename="wlc_cert_errors.log"')
                msg.attach(error_attachment)
        
        # Send email
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as smtp:
            smtp.starttls()
            smtp.login(SENDER_EMAIL, SENDER_PASSWORD)
            smtp.send_message(msg)
        
        logger.info(f"Email sent successfully to {', '.join(send_to)}")
        return True
        
    except Exception as e:
        logger.error(f"Error sending email: {str(e)}")
        error_logger.error(f"Email Send Error: {str(e)}")
        return False


def process_all_wlcs(send_email_report=True, email_recipients=None, max_threads=MAX_THREADS):
    """Process all WLCs and generate certificate report"""
    try:
        start_time = time.time()
        
        global results
        results = []
        
        with open(WLC_DATA_FILE, 'r') as f:
            wlc_devices = json.load(f)
        
        logger.info(f"Loaded {len(wlc_devices)} WLC devices from {WLC_DATA_FILE}")
        
        secure_session = SecureWLCSession(session_timeout=1800)
        
        if not secure_session.authenticate():
            logger.error("Authentication failed")
            return False
        
        commands = ["show crypto pki certificates"]
        
        now_tz = datetime.now(TZ)
        timestamp = now_tz.strftime('%d %b %Y %H%M')
        output_file = f"WLC_Certificate_Status_Report_{timestamp}.xlsx"
        
        total_wlcs = len(wlc_devices)
        logger.info(f"Processing {total_wlcs} WLCs using {max_threads} concurrent threads")
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_threads) as executor:
            future_to_device = {
                executor.submit(process_wlc, device, secure_session, commands, i + 1): device 
                for i, device in enumerate(wlc_devices)
            }
            
            completed = 0
            for future in concurrent.futures.as_completed(future_to_device):
                completed += 1
                progress = (completed / total_wlcs) * 100
                logger.info(f"Progress: {completed}/{total_wlcs} ({progress:.1f}%)")
                
                try:
                    result = future.result()
                    device_name = future_to_device[future]['HOSTNAME']
                    if result['success']:
                        logger.info(f"[OK] {device_name}")
                    else:
                        logger.warning(f"[FAILED] {device_name}")
                except Exception as e:
                    device_name = future_to_device[future]['HOSTNAME']
                    logger.error(f"Thread error for {device_name}: {str(e)}")
                    error_logger.error(f"Thread error for {device_name}: {str(e)}")
        
        logger.info("Processing certificate data...")
        
        certificate_data = []
        success_count = 0
        failed_count = 0
        failed_wlcs = []
        total_certs = 0
        expired_certs = 0
        expiring_soon = 0
        
        for result in results:
            if result['success']:
                success_count += 1
                
                certs = parse_certificates(result['output'])
                total_certs += len(certs)
                
                for cert in certs:
                    remarks, expiry_date = calculate_cert_remarks(cert['end_date'])
                    
                    if remarks == 'EXPIRED':
                        expired_certs += 1
                    elif 'remaining' in remarks:
                        expiring_soon += 1
                    
                    cert_entry = {
                        'region': result['region'],
                        'site_id': result['site_id'],
                        'hostname': result['hostname'],
                        'ip': result['ip'],
                        'cert_type': cert['cert_type'],
                        'name': cert['name'],
                        'start_date': cert['start_date'],
                        'end_date': cert['end_date'],
                        'trustpoints': cert['trustpoints'],
                        'status': cert['status'],
                        'storage': cert['storage'],
                        'remarks': remarks
                    }
                    
                    certificate_data.append(cert_entry)
            else:
                failed_count += 1
                failed_wlcs.append({
                    'hostname': result['hostname'],
                    'ip': result['ip'],
                    'reason': result['output']
                })
        
        if certificate_data:
            create_excel_report(certificate_data, output_file)
            logger.info(f"Certificate report created: {output_file}")
        else:
            logger.warning("No certificate data to report")
            return False
        
        end_time = time.time()
        execution_time = end_time - start_time
        success_rate = (success_count / max(total_wlcs, 1)) * 100
        
        print("\n" + "="*60)
        print("EXECUTION SUMMARY")
        print("="*60)
        logger.info(f"Certificate report saved: {output_file}")
        logger.info(f"Total WLCs: {total_wlcs}")
        logger.info(f"Successful: {success_count}/{total_wlcs} ({success_rate:.1f}%)")
        logger.info(f"Failed: {failed_count}/{total_wlcs}")
        logger.info(f"Total certificates: {total_certs}")
        logger.info(f"Expired certificates: {expired_certs}")
        logger.info(f"Expiring soon (< 1 year): {expiring_soon}")
        logger.info(f"Execution time: {execution_time:.2f}s ({execution_time/60:.1f}m)")
        print("="*60 + "\n")
        
        summary_data = {
            'total_wlcs': total_wlcs,
            'successful': success_count,
            'failed': failed_count,
            'total_certs': total_certs,
            'expired_certs': expired_certs,
            'expiring_soon': expiring_soon,
            'execution_time': execution_time,
            'failures': failed_wlcs
        }
        
        if send_email_report:
            logger.info("Sending email report...")
            now_tz = datetime.now(TZ)
            subject = f"WLC Certificate Status Report - {now_tz.strftime('%d %b %Y').upper()}"
            email_success = send_email(output_file, subject, email_recipients, summary_data)
            
            if email_success:
                logger.info("Email sent successfully")
            else:
                logger.error("Failed to send email")
        
        logger.info("Cleaning up secure session...")
        del secure_session
        gc.collect()
        
        return True
        
    except FileNotFoundError:
        logger.error(f"Error: {WLC_DATA_FILE} not found")
        return False
    except json.JSONDecodeError:
        logger.error(f"Error: {WLC_DATA_FILE} is not valid JSON")
        error_logger.error(f"JSON decode error in {WLC_DATA_FILE}")
        return False
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        error_logger.error(f"Unexpected error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def validate_environment():
    """Validate required environment variables"""
    required_vars = []
    
    if not SENDER_EMAIL:
        required_vars.append('SENDER_EMAIL')
    if not SENDER_PASSWORD:
        required_vars.append('SENDER_PASSWORD')
    
    if required_vars:
        print("Missing required environment variables:")
        for var in required_vars:
            print(f"   - {var}")
        print("\nPlease check your .env file")
        return False
    
    return True


def create_sample_json():
    """Create sample WLC JSON file"""
    sample_devices = [
        {
            "SITE ID": "SITE01",
            "HOSTNAME": "WLC-SITE01",
            "IP ADDRESS": "192.168.1.10",
            "MODEL": "C9800-80-K9",
            "REGION": "Americas",
            "CITY": "New York",
            "TIMEZONE": "UTC-5"
        },
        {
            "SITE ID": "SITE02",
            "HOSTNAME": "WLC-SITE02",
            "IP ADDRESS": "192.168.1.11",
            "MODEL": "C9800-L-K9",
            "REGION": "EMEA",
            "CITY": "London",
            "TIMEZONE": "UTC+0"
        },
        {
            "SITE ID": "SITE03",
            "HOSTNAME": "WLC-SITE03",
            "IP ADDRESS": "192.168.1.12",
            "MODEL": "C9800-CL-K9",
            "REGION": "APAC",
            "CITY": "Singapore",
            "TIMEZONE": "UTC+8"
        }
    ]
    
    with open(WLC_DATA_FILE, 'w') as f:
        json.dump(sample_devices, f, indent=4)
    
    logger.info(f"Sample JSON created: {WLC_DATA_FILE}")
    print(f"Sample JSON created: {WLC_DATA_FILE}")
    print("Update with your actual WLC information before running")


def create_env_file():
    """Create sample .env configuration file"""
    env_content = """# Cisco 9800 WLC Certificate Status Reporter Configuration
# ============================================================

# Email Configuration (REQUIRED for email reports)
# =================================================
SENDER_EMAIL=your-email@example.com
SENDER_PASSWORD=your-app-password-here
DEFAULT_RECIPIENTS=recipient1@example.com,recipient2@example.com

# SMTP Configuration (defaults shown)
# ====================================
SMTP_SERVER=smtp.gmail.com
SMTP_PORT=587

# File Configuration
# ==================
WLC_DATA_FILE=wlc.json
LOG_DIR=logs

# Processing Configuration
# ========================
MAX_THREADS=10
COMMAND_TIMEOUT=180
CONNECTION_TIMEOUT=180
PING_TIMEOUT=5
MAX_RETRIES=3

# Report Configuration
# ====================
REPORT_TIMEZONE=UTC

# Security Notes:
# ==============
# 1. Set secure permissions: chmod 600 .env
# 2. Add .env to .gitignore
# 3. Never commit this file
# 4. Use app passwords for email
# 5. Rotate credentials regularly
"""
    
    with open('.env.example', 'w') as f:
        f.write(env_content)
    
    logger.info("Sample .env file created: .env.example")
    print("Sample .env file created: .env.example")
    print("\nSetup steps:")
    print("1. Copy .env.example to .env")
    print("2. Update .env with your credentials")
    print("3. Set permissions: chmod 600 .env")


def test_environment():
    """Test environment configuration"""
    print("\n" + "="*60)
    print("Testing Environment Configuration")
    print("="*60 + "\n")
    
    missing_vars = []
    
    if not SENDER_EMAIL:
        missing_vars.append('SENDER_EMAIL')
    if not SENDER_PASSWORD:
        missing_vars.append('SENDER_PASSWORD')
    
    if missing_vars:
        print("Missing required variables:")
        for var in missing_vars:
            print(f"   - {var}")
        return False
    
    print("Environment variables found:")
    print(f"   SENDER_EMAIL: {SENDER_EMAIL}")
    print(f"   SENDER_PASSWORD: {'*' * len(SENDER_PASSWORD)}")
    
    if DEFAULT_RECIPIENTS:
        print(f"   DEFAULT_RECIPIENTS: {', '.join(DEFAULT_RECIPIENTS)}")
    
    print(f"\nConfiguration:")
    print(f"   SMTP_SERVER: {SMTP_SERVER}")
    print(f"   SMTP_PORT: {SMTP_PORT}")
    print(f"   MAX_THREADS: {MAX_THREADS}")
    print(f"   COMMAND_TIMEOUT: {COMMAND_TIMEOUT}s")
    print(f"   WLC_DATA_FILE: {WLC_DATA_FILE}")
    print(f"   REPORT_TIMEZONE: {TIMEZONE}")
    
    print("\nEnvironment configuration valid")
    return True


def main():
    """Main function"""
    
    print("="*80)
    print("Cisco 9800 Series WLC Certificate Status Reporter")
    print("="*80)
    print("\nSecurity Features:")
    print("   - AES-128-CBC encryption for passwords")
    print("   - PBKDF2-SHA256 key derivation (100k iterations)")
    print("   - Session-based credential management")
    print("   - Automatic memory cleanup\n")
    
    with open(ERROR_LOG_FILE, 'w') as f:
        f.write(f"WLC Certificate Status Error Log - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("="*60 + "\n\n")
    
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Cisco 9800 WLC Certificate Status Reporter',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python run_job.py                          # Run with defaults
  python run_job.py --create-json            # Create sample JSON
  python run_job.py --create-env             # Create sample .env
  python run_job.py --test-env               # Test configuration
  python run_job.py --no-email               # Run without email
  python run_job.py --threads 5              # Use 5 threads
  python run_job.py --email user@example.com # Custom recipients

Setup:
  1. python run_job.py --create-env
  2. cp .env.example .env
  3. Edit .env with your credentials
  4. python run_job.py --create-json
  5. Edit wlc.json with your WLCs
  6. python run_job.py
        """
    )
    
    parser.add_argument('--create-json', action='store_true', 
                       help='Create sample JSON file')
    parser.add_argument('--create-env', action='store_true', 
                       help='Create sample .env file')
    parser.add_argument('--test-env', action='store_true', 
                       help='Test environment configuration')
    parser.add_argument('--no-email', action='store_true', 
                       help='Do not send email report')
    parser.add_argument('--email', type=str, nargs='+', 
                       help='Email recipients (space-separated)')
    parser.add_argument('--threads', type=int, default=MAX_THREADS, 
                       help=f'Concurrent threads (default: {MAX_THREADS})')
    
    args = parser.parse_args()
    
    if args.create_json:
        create_sample_json()
    elif args.create_env:
        create_env_file()
    elif args.test_env:
        test_environment()
    else:
        if not validate_environment():
            print("\nEnvironment validation failed")
            print("Use --create-env to create sample configuration")
            return 1
        
        print("Environment validated")
        print(f"Email: {SENDER_EMAIL}")
        print(f"Recipients: {', '.join(DEFAULT_RECIPIENTS) if DEFAULT_RECIPIENTS else 'None'}\n")
        
        if not os.path.isfile(WLC_DATA_FILE):
            print(f"Error: {WLC_DATA_FILE} not found")
            print("Use --create-json to create sample file")
            return 1
        
        success = process_all_wlcs(
            send_email_report=not args.no_email,
            email_recipients=args.email,
            max_threads=args.threads
        )
        
        if success:
            print("WLC Certificate Status check completed successfully\n")
            return 0
        else:
            print("WLC Certificate Status check failed\n")
            return 1


if __name__ == "__main__":
    sys.exit(main())